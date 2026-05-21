from fastapi import FastAPI, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import Session
from datetime import timedelta
from typing import List, Optional
import shutil
from pathlib import Path
import io
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from .db.session import engine, get_db, Base
from .models import models
from .schemas import schemas
from .core import auth
import os
from dotenv import load_dotenv

load_dotenv()

# Setup upload directory
UPLOAD_DIR = Path("uploads")
UPLOAD_DIR.mkdir(exist_ok=True)

# Create tables
models.Base.metadata.create_all(bind=engine)

app = FastAPI(title="Grupo Ipurre EIRL - Sistema de Gestión de Obras API")

# Configure CORS
origins = [
    "http://localhost:5173",
    "http://127.0.0.1:5173",
    "http://localhost:3000",
    "http://127.0.0.1:3000",
    "http://localhost:8080",
    "http://127.0.0.1:8080",
    "http://localhost:5174",
    "http://127.0.0.1:5174",
    "http://localhost:4173",
    "http://127.0.0.1:4173",
]

app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

oauth2_scheme = OAuth2PasswordBearer(tokenUrl="token")

@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    file_path = UPLOAD_DIR / file.filename
    with file_path.open("wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
    return {"url": f"/uploads/{file.filename}", "filename": file.filename}

@app.get("/uploads/{filename}")
async def get_uploaded_file(filename: str):
    file_path = UPLOAD_DIR / filename
    if not file_path.exists():
        raise HTTPException(status_code=404, detail="Archivo no encontrado")
    return FileResponse(file_path)

@app.post("/token", response_model=schemas.Token)
async def login_for_access_token(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.email == form_data.username).first()
    if not user or not auth.verify_password(form_data.password, user.password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    access_token_expires = timedelta(minutes=auth.ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = auth.create_access_token(
        data={"sub": user.email}, expires_delta=access_token_expires
    )
    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/users/me", response_model=schemas.User)
async def read_users_me(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    try:
        payload = auth.jwt.decode(token, auth.SECRET_KEY, algorithms=[auth.ALGORITHM])
        email: str = payload.get("sub")
        if email is None:
            raise HTTPException(status_code=401, detail="Invalid token")
    except auth.jwt.JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    user = db.query(models.User).filter(models.User.email == email).first()
    if user is None:
        raise HTTPException(status_code=404, detail="User not found")
    return user

@app.get("/proyectos", response_model=List[schemas.Project])
async def get_proyectos(db: Session = Depends(get_db)):
    proyectos = db.query(models.Project).all()
    return proyectos

@app.post("/proyectos", response_model=schemas.Project)
async def create_proyecto(proyecto: schemas.ProjectCreate, db: Session = Depends(get_db)):
    db_proyecto = models.Project(**proyecto.model_dump())
    db.add(db_proyecto)
    
    # Create notification for project creation
    notif = models.Notification(
        usuario_id=proyecto.responsable_id,
        proyecto_id=None, # Or assign the new ID after refresh
        tipo="Info",
        titulo="Nuevo Proyecto Asignado",
        mensaje=f"Se te ha asignado como responsable del proyecto: {proyecto.nombre}",
        url_enlace="/proyectos"
    )
    db.add(notif)
    
    db.commit()
    db.refresh(db_proyecto)
    return db_proyecto

@app.get("/proyectos/{proyecto_id}", response_model=schemas.Project)
async def get_proyecto(proyecto_id: int, db: Session = Depends(get_db)):
    proyecto = db.query(models.Project).filter(models.Project.id == proyecto_id).first()
    if not proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    return proyecto

@app.put("/proyectos/{proyecto_id}", response_model=schemas.Project)
async def update_proyecto(proyecto_id: int, proyecto_update: schemas.ProjectBase, db: Session = Depends(get_db)):
    db_proyecto = db.query(models.Project).filter(models.Project.id == proyecto_id).first()
    if not db_proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    
    update_data = proyecto_update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_proyecto, key, value)
    
    db.commit()
    db.refresh(db_proyecto)
    return db_proyecto

@app.delete("/proyectos/{proyecto_id}")
async def delete_proyecto(proyecto_id: int, db: Session = Depends(get_db)):
    db_proyecto = db.query(models.Project).filter(models.Project.id == proyecto_id).first()
    if not db_proyecto:
        raise HTTPException(status_code=404, detail="Proyecto no encontrado")
    db.delete(db_proyecto)
    db.commit()
    return {"message": "Proyecto eliminado"}

# Endpoints para Actividades
@app.get("/proyectos/{proyecto_id}/actividades", response_model=List[schemas.Actividad])
async def get_actividades(proyecto_id: int, db: Session = Depends(get_db)):
    return db.query(models.Actividad).filter(models.Actividad.proyecto_id == proyecto_id).all()

@app.post("/proyectos/{proyecto_id}/actividades", response_model=schemas.Actividad)
async def create_actividad(proyecto_id: int, actividad: schemas.ActividadCreate, db: Session = Depends(get_db)):
    db_actividad = models.Actividad(**actividad.model_dump())
    db_actividad.proyecto_id = proyecto_id
    db.add(db_actividad)
    db.commit()
    db.refresh(db_actividad)
    return db_actividad

# Endpoints para Planos
@app.get("/proyectos/{proyecto_id}/planos", response_model=List[schemas.Plano])
async def get_planos(proyecto_id: int, db: Session = Depends(get_db)):
    return db.query(models.Plano).filter(models.Plano.proyecto_id == proyecto_id).all()

@app.post("/proyectos/{proyecto_id}/planos", response_model=schemas.Plano)
async def create_plano(proyecto_id: int, plano: schemas.PlanoCreate, db: Session = Depends(get_db)):
    db_plano = models.Plano(**plano.model_dump())
    db_plano.proyecto_id = proyecto_id
    db.add(db_plano)
    db.commit()
    db.refresh(db_plano)
    return db_plano

@app.get("/planos/{plano_id}/versiones", response_model=List[schemas.VersionPlano])
async def get_plano_versiones(plano_id: int, db: Session = Depends(get_db)):
    return db.query(models.VersionPlano).filter(models.VersionPlano.plano_id == plano_id).all()

@app.post("/planos/{plano_id}/versiones", response_model=schemas.VersionPlano)
async def create_plano_version(plano_id: int, version: schemas.VersionPlanoCreate, db: Session = Depends(get_db)):
    db_version = models.VersionPlano(**version.model_dump())
    db_version.plano_id = plano_id
    db.add(db_version)
    db.commit()
    db.refresh(db_version)
    return db_version

@app.delete("/planos/{plano_id}/versiones/{version_id}")
async def delete_plano_version(plano_id: int, version_id: int, db: Session = Depends(get_db)):
    db_version = db.query(models.VersionPlano).filter(
        models.VersionPlano.id == version_id,
        models.VersionPlano.plano_id == plano_id
    ).first()
    if not db_version:
        raise HTTPException(status_code=404, detail="Versión no encontrada")
    db.delete(db_version)
    db.commit()
    return {"message": "Versión eliminada"}

@app.delete("/planos/{plano_id}")
async def delete_plano(plano_id: int, db: Session = Depends(get_db)):
    db_plano = db.query(models.Plano).filter(models.Plano.id == plano_id).first()
    if not db_plano:
        raise HTTPException(status_code=404, detail="Plano no encontrado")
    db.delete(db_plano)
    db.commit()
    return {"message": "Plano eliminado"}

@app.get("/planos/{plano_id}", response_model=schemas.Plano)
async def get_plano(plano_id: int, db: Session = Depends(get_db)):
    plano = db.query(models.Plano).filter(models.Plano.id == plano_id).first()
    if not plano:
        raise HTTPException(status_code=404, detail="Plano no encontrado")
    return plano

@app.get("/planos/{plano_id}/proyecto", response_model=schemas.Project)
async def get_plano_proyecto(plano_id: int, db: Session = Depends(get_db)):
    plano = db.query(models.Plano).filter(models.Plano.id == plano_id).first()
    if not plano:
        raise HTTPException(status_code=404, detail="Plano no encontrado")
    proyecto = db.query(models.Project).filter(models.Project.id == plano.proyecto_id).first()
    return proyecto

@app.delete("/actividades/{actividad_id}")
async def delete_actividad(actividad_id: int, db: Session = Depends(get_db)):
    db_actividad = db.query(models.Actividad).filter(models.Actividad.id == actividad_id).first()
    if not db_actividad:
        raise HTTPException(status_code=404, detail="Actividad no encontrada")
    db.delete(db_actividad)
    db.commit()
    return {"message": "Actividad eliminada"}

@app.put("/actividades/{actividad_id}", response_model=schemas.Actividad)
async def update_actividad(actividad_id: int, actividad_update: schemas.ActividadBase, db: Session = Depends(get_db)):
    db_actividad = db.query(models.Actividad).filter(models.Actividad.id == actividad_id).first()
    if not db_actividad:
        raise HTTPException(status_code=404, detail="Actividad no encontrada")
    
    update_data = actividad_update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_actividad, key, value)
    
    db.commit()
    db.refresh(db_actividad)
    return db_actividad

@app.delete("/evidencias/{evidencia_id}")
async def delete_evidencia(evidencia_id: int, db: Session = Depends(get_db)):
    db_evidencia = db.query(models.Evidencia).filter(models.Evidencia.id == evidencia_id).first()
    if not db_evidencia:
        raise HTTPException(status_code=404, detail="Evidencia no encontrada")
    db.delete(db_evidencia)
    db.commit()
    return {"message": "Evidencia eliminada"}

# Endpoints para Evidencias
@app.get("/evidencias", response_model=List[schemas.Evidencia])
async def get_all_evidencias(db: Session = Depends(get_db)):
    return db.query(models.Evidencia).all()

@app.get("/proyectos/{proyecto_id}/evidencias", response_model=List[schemas.Evidencia])
async def get_evidencias(proyecto_id: int, db: Session = Depends(get_db)):
    return db.query(models.Evidencia).filter(models.Evidencia.proyecto_id == proyecto_id).all()

@app.post("/proyectos/{proyecto_id}/evidencias", response_model=schemas.Evidencia)
async def create_evidencia(proyecto_id: int, evidencia: schemas.EvidenciaCreate, db: Session = Depends(get_db)):
    db_evidencia = models.Evidencia(**evidencia.model_dump())
    db_evidencia.proyecto_id = proyecto_id
    db.add(db_evidencia)
    db.commit()
    db.refresh(db_evidencia)
    return db_evidencia

@app.post("/evidencias-general", response_model=schemas.Evidencia)
async def create_evidencia_general(evidencia: schemas.EvidenciaCreate, db: Session = Depends(get_db)):
    db_evidencia = models.Evidencia(**evidencia.model_dump())
    db.add(db_evidencia)
    db.commit()
    db.refresh(db_evidencia)
    return db_evidencia

# Endpoints para Usuarios
@app.get("/users", response_model=List[schemas.User])
async def get_users(db: Session = Depends(get_db)):
    return db.query(models.User).all()

@app.post("/users", response_model=schemas.User)
async def create_user(user: schemas.UserCreate, db: Session = Depends(get_db)):
    hashed_password = auth.get_password_hash(user.password)
    db_user = models.User(
        nombre=user.nombre,
        email=user.email,
        telefono=user.telefono,
        password=hashed_password,
        rol_id=user.rol_id
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user

@app.get("/users/{user_id}", response_model=schemas.User)
async def get_user(user_id: int, db: Session = Depends(get_db)):
    user = db.query(models.User).filter(models.User.id == user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    return user

@app.put("/users/{user_id}", response_model=schemas.User)
async def update_user(user_id: int, user_update: schemas.UserBase, db: Session = Depends(get_db)):
    db_user = db.query(models.User).filter(models.User.id == user_id).first()
    if not db_user:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    update_data = user_update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(db_user, key, value)
    
    db.commit()
    db.refresh(db_user)
    return db_user

@app.get("/roles", response_model=List[schemas.Role])
async def get_roles(db: Session = Depends(get_db)):
    return db.query(models.Role).all()

# Endpoints para Notificaciones
@app.get("/notifications", response_model=List[schemas.Notification])
async def get_notifications(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    payload = auth.jwt.decode(token, auth.SECRET_KEY, algorithms=[auth.ALGORITHM])
    email: str = payload.get("sub")
    user = db.query(models.User).filter(models.User.email == email).first()
    return db.query(models.Notification).filter(models.Notification.usuario_id == user.id).order_by(models.Notification.fecha_creacion.desc()).all()

@app.post("/notifications", response_model=schemas.Notification)
async def create_notification(notif: schemas.NotificationCreate, db: Session = Depends(get_db)):
    db_notif = models.Notification(**notif.model_dump())
    db.add(db_notif)
    db.commit()
    db.refresh(db_notif)
    return db_notif

@app.get("/reports/stats", response_model=schemas.DashboardStats)
async def get_report_stats(db: Session = Depends(get_db)):
    proyectos = db.query(models.Project).all()
    total = len(proyectos)
    activos = len([p for p in proyectos if p.estado == 'Activo'])
    presupuesto = sum([p.presupuesto or 0 for p in proyectos])
    avance_avg = sum([p.avance_total or 0 for p in proyectos]) / total if total > 0 else 0
    
    # Mock financial data (in a real app, this would come from an expenses table)
    financials = [
        {"month": "Ene", "presupuesto": 400000, "gasto": 380000, "avance": 40},
        {"month": "Feb", "presupuesto": 450000, "gasto": 460000, "avance": 45},
        {"month": "Mar", "presupuesto": 550000, "gasto": 510000, "avance": 55},
        {"month": "Abr", "presupuesto": 600000, "gasto": 580000, "avance": 60},
    ]
    
    # Distribution by type
    dist = {
        "Residencial": len([p for p in proyectos if p.tipo == 'Residencial']),
        "Comercial": len([p for p in proyectos if p.tipo == 'Comercial']),
        "Industrial": len([p for p in proyectos if p.tipo == 'Industrial']),
    }
    
    total_dist = sum(dist.values())
    distribution = []
    for name, count in dist.items():
        val = (count / total_dist * 100) if total_dist > 0 else 0
        distribution.append({"name": name, "value": int(val)})
    
    return {
        "summary": {
            "total_proyectos": total,
            "proyectos_activos": activos,
            "presupuesto_total": presupuesto,
            "gasto_estimado": int(presupuesto * 0.95), # Mock
            "avance_promedio": avance_avg
        },
        "financials": financials,
        "distribution": distribution
    }


@app.get("/reports/export/excel")
async def export_report_excel(db: Session = Depends(get_db)):
    # Get all projects data
    proyectos = db.query(models.Project).all()
    
    # Create workbook and select active worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Proyectos"
    
    # Define headers
    headers = [
        "ID", "Nombre", "Descripción", "Ubicación", "Cliente", 
        "Fecha Inicio", "Fecha Fin", "Presupuesto", "Avance Total (%)", 
        "Estado", "Tipo", "Responsable ID"
    ]
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Add project data
    for row, proyecto in enumerate(proyectos, 2):
        ws.cell(row=row, column=1, value=proyecto.id)
        ws.cell(row=row, column=2, value=proyecto.nombre)
        ws.cell(row=row, column=3, value=proyecto.descripcion)
        ws.cell(row=row, column=4, value=proyecto.ubicacion)
        ws.cell(row=row, column=5, value=proyecto.cliente)
        ws.cell(row=row, column=6, value=str(proyecto.fecha_inicio) if proyecto.fecha_inicio else "")
        ws.cell(row=row, column=7, value=str(proyecto.fecha_fin) if proyecto.fecha_fin else "")
        ws.cell(row=row, column=8, value=proyecto.presupuesto or 0)
        ws.cell(row=row, column=9, value=proyecto.avance_total or 0)
        ws.cell(row=row, column=10, value=proyecto.estado)
        ws.cell(row=row, column=11, value=proyecto.tipo)
        ws.cell(row=row, column=12, value=proyecto.responsable_id)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Return as streaming response
    return StreamingResponse(
        io.BytesIO(buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte_proyectos.xlsx"}
    )


@app.get("/reports/export/pdf")
async def export_report_pdf(db: Session = Depends(get_db)):
    # Get all projects data
    proyectos = db.query(models.Project).all()
    
    # Create buffer
    buffer = io.BytesIO()
    
    # Create PDF document
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Define styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Add title
    title = Paragraph("Reporte de Proyectos - Grupo Ipurre EIRL", title_style)
    elements.append(title)
    elements.append(Spacer(1, 12))
    
    # Prepare table data
    data = [["ID", "Nombre", "Cliente", "Presupuesto", "Avance (%)", "Estado", "Tipo"]]
    
    for proyecto in proyectos:
        data.append([
            str(proyecto.id),
            proyecto.nombre,
            proyecto.cliente or "",
            f"${proyecto.presupuesto:,.2f}" if proyecto.presupuesto else "$0.00",
            f"{proyecto.avance_total or 0}%",
            proyecto.estado,
            proyecto.tipo
        ])
    
    # Create table
    table = Table(data)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    # Return as streaming response
    buffer.seek(0)
    return StreamingResponse(
        io.BytesIO(buffer.read()),
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=reporte_proyectos.pdf"}
    )

@app.put("/notifications/{notification_id}/read")
async def mark_notification_as_read(notification_id: int, db: Session = Depends(get_db)):
    db_notif = db.query(models.Notification).filter(models.Notification.id == notification_id).first()
    if not db_notif:
        raise HTTPException(status_code=404, detail="Notificación no encontrada")
    db_notif.leida = True
    db.commit()
    return {"message": "Notificación marcada como leída"}

# Endpoints para Documentos
@app.get("/documents", response_model=List[schemas.Document])
async def get_all_documents(db: Session = Depends(get_db)):
    return db.query(models.Document).all()

@app.get("/documents/{document_id}", response_model=schemas.Document)
async def get_document(document_id: int, db: Session = Depends(get_db)):
    doc = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    return doc

@app.post("/documents", response_model=schemas.Document)
async def create_document(document: schemas.DocumentCreate, db: Session = Depends(get_db)):
    db_doc = models.Document(**document.model_dump())
    db.add(db_doc)
    db.commit()
    db.refresh(db_doc)
    return db_doc

@app.delete("/documents/{document_id}")
async def delete_document(document_id: int, db: Session = Depends(get_db)):
    doc = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    db.delete(doc)
    db.commit()
    return {"message": "Documento eliminado"}

@app.put("/documents/{document_id}", response_model=schemas.Document)
async def update_document(document_id: int, document_update: schemas.DocumentUpdate, db: Session = Depends(get_db)):
    doc = db.query(models.Document).filter(models.Document.id == document_id).first()
    if not doc:
        raise HTTPException(status_code=404, detail="Documento no encontrado")
    update_data = document_update.model_dump(exclude_unset=True)
    for key, value in update_data.items():
        setattr(doc, key, value)
    db.commit()
    db.refresh(doc)
    return doc
